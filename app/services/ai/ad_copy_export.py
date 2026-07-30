"""Export generated ad copy to Excel / CSV / JSON.

Operates on a persisted :class:`AdCopyGeneration` row. Excel is the primary
deliverable (the user pastes it straight into Google Ads), with a sheet per
asset type; CSV flattens headlines+descriptions for a quick copy-paste.
"""

from __future__ import annotations

import csv
import io
import json
from typing import Any

from app.models.ad_copy import AdCopyGeneration

_H = "Text"
_L = "Chars"
_R = "Reason"
_P = "Pin"


def _assets(gen: AdCopyGeneration) -> dict[str, Any]:
    return gen.generated_assets or {}


def render_json(gen: AdCopyGeneration) -> str:
    payload = {
        "campus": gen.campus,
        "final_url": gen.final_url,
        "backend": gen.backend,
        "generated_at": gen.created_at.isoformat() if gen.created_at else None,
        "assets": gen.generated_assets,
        "keywords": (gen.keyword_snapshot or {}).get("keywords", []),
        "scores": gen.scores,
        "reasoning": gen.reasoning,
    }
    return json.dumps(payload, indent=2, default=str)


def render_csv(gen: AdCopyGeneration) -> str:
    a = _assets(gen)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Asset Type", _H, _L, _R])
    for h in a.get("headlines", []):
        w.writerow(["Headline", h.get("text"), h.get("length"), h.get("reason")])
    for d in a.get("descriptions", []):
        w.writerow(["Description", d.get("text"), d.get("length"), d.get("reason")])
    for p in a.get("display_paths", []):
        w.writerow(["Display Path", p, len(p), ""])
    for c in a.get("callouts", []):
        w.writerow(["Callout", c, len(c), ""])
    for n in a.get("negative_keywords", []):
        w.writerow(["Negative Keyword", n, "", ""])
    return buf.getvalue()


def render_excel(gen: AdCopyGeneration) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export.") from exc

    a = _assets(gen)
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1E40AF")
    head_font = Font(bold=True, color="FFFFFF")

    def _header(ws, cols: list[str]) -> None:
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font

    # Summary
    ws = wb.active
    ws.title = "Summary"
    _header(ws, ["Field", "Value"])
    q = (gen.scores or {}).get("quality", {})
    for k, v in [
        ("Campus", gen.campus),
        ("Final URL", gen.final_url),
        ("URL source", gen.url_source),
        ("URL confidence", float(gen.url_confidence) if gen.url_confidence is not None else None),
        ("Backend", gen.backend),
        ("Expected Ad Strength", q.get("expected_ad_strength")),
        ("Headlines", q.get("headline_count")),
        ("Descriptions", q.get("description_count")),
        ("Keyword coverage", q.get("keyword_coverage")),
        ("Generated at", gen.created_at.isoformat() if gen.created_at else None),
    ]:
        ws.append([k, v])

    # Headlines
    hs = wb.create_sheet("Headlines")
    _header(hs, ["#", _H, _L, _P, _R])
    for i, h in enumerate(a.get("headlines", []), 1):
        hs.append([i, h.get("text"), h.get("length"), h.get("pinned_position"), h.get("reason")])

    # Descriptions
    ds = wb.create_sheet("Descriptions")
    _header(ds, ["#", _H, _L, _R])
    for i, d in enumerate(a.get("descriptions", []), 1):
        ds.append([i, d.get("text"), d.get("length"), d.get("reason")])

    # Extensions
    es = wb.create_sheet("Extensions")
    _header(es, ["Type", "Value"])
    for p in a.get("display_paths", []):
        es.append(["Display Path", p])
    for c in a.get("callouts", []):
        es.append(["Callout", c])
    for label, vals in (a.get("structured_snippets") or {}).items():
        es.append([f"Snippet: {label}", ", ".join(vals)])
    for s in a.get("sitelinks", []):
        es.append(["Sitelink", s.get("text")])

    # Keywords (scored intelligence)
    ks = wb.create_sheet("Keywords")
    _header(ks, ["Keyword", "Intent", "Score", "Source", "Clicks", "CTR", "CPC", "QS",
                 "Match Type", "Why match type",
                 "Suggested Bid (max CPC)", "Bid Basis", "Why this bid"])
    for kw in (gen.keyword_snapshot or {}).get("keywords", []):
        ks.append([
            kw.get("keyword"), kw.get("intent"), kw.get("score"), kw.get("source"),
            kw.get("historical_clicks"), kw.get("historical_ctr"),
            kw.get("historical_cpc"), kw.get("quality_score"),
            kw.get("recommended_match_type"), kw.get("match_reason"),
            kw.get("recommended_bid"), kw.get("bid_basis"), kw.get("bid_reason"),
        ])

    # Campaign Keywords (paste-ready, grouped by ad group + match type)
    ck = wb.create_sheet("Campaign Keywords")
    _header(ck, ["Ad Group", "Keyword (paste into Google Ads)", "Match Types", "Suggested Bid"])
    for grp in (gen.keyword_snapshot or {}).get("groups", []):
        match_types = ", ".join(grp.get("recommended_match_types", []))
        bid = grp.get("recommended_bid")
        for kw in grp.get("match_keywords", []):
            ck.append([grp.get("name"), kw, match_types, bid])

    # ---- Campaign plan sheets (only when a budget was provided) ----
    plan = (gen.scores or {}).get("campaign_plan") or {}
    seasonality = (gen.scores or {}).get("seasonality") or {}
    if plan.get("available"):
        f = plan.get("forecast") or {}
        est = " (ESTIMATE)" if f.get("cpl_is_estimated") else ""

        bp = wb.create_sheet("Budget Plan")
        _header(bp, ["Ad Group", "Phase", "Budget (INR)", "Avg CPC", "Est. Clicks",
                     "Est. Impressions", "Est. Leads" + est, "Est. CPL" + est, "Bidding"])
        for r in plan.get("allocation", []):
            bp.append([r.get("ad_group"), r.get("phase"), r.get("budget"), r.get("avg_cpc"),
                       r.get("est_clicks"), r.get("est_impressions"), r.get("est_leads"),
                       r.get("est_cpl"), r.get("bidding")])
        bp.append([])
        bp.append(["TOTAL", "", f.get("budget"), f.get("blended_cpc"), f.get("est_clicks"),
                   f.get("est_impressions"), f.get("est_leads"), f.get("est_cpl"), ""])
        rl = plan.get("realism") or {}
        if rl:
            bp.append([])
            bp.append(["REALITY CHECK — clicks don't scale linearly with budget"])
            bp.append(["Realistic clicks:",
                       f"{rl.get('realistic_clicks_low'):,}–{rl.get('realistic_clicks_high'):,}"
                       f"  (flat-CPC optimistic: {rl.get('arithmetic_clicks'):,})"])
            bp.append(["Effective CPC at scale:", f"₹{rl.get('effective_cpc')} "
                       f"(historical ₹{rl.get('hist_cpc')})"])
            bp.append(["Your real history:",
                       f"{rl.get('hist_clicks_per_year'):,} clicks/yr @ "
                       f"₹{rl.get('hist_spend_per_year'):,}/yr"])
            if rl.get("budget_multiple"):
                bp.append(["Budget vs history:", f"{rl.get('budget_multiple')}×"])
            if rl.get("annual_search_demand"):
                bp.append(["Annual search demand:", f"{rl.get('annual_search_demand'):,}"])
            if rl.get("click_ceiling"):
                bp.append(["Absolute click ceiling:", f"{rl.get('click_ceiling'):,}"])
            bp.append(["Note:", rl.get("note")])

        bp.append([])
        bp.append([f"Leads/CPL assume a {round((f.get('assumed_cvr') or 0) * 100, 1)}% "
                   "click→lead conversion rate (no conversion tracking on this account yet)."])
        bid = plan.get("bidding") or {}
        bp.append([])
        bp.append(["Recommended bidding:", bid.get("recommended") or bid.get("primary")])
        if bid.get("why"):
            bp.append(["Why:", bid.get("why")])
        if bid.get("daily_budget"):
            bp.append(["Daily budget:", f"₹{bid.get('daily_budget')}/day"])
        if bid.get("max_cpc_cap"):
            bp.append(["Max-CPC cap:", f"₹{bid.get('max_cpc_cap')}"])
        if bid.get("options"):
            bp.append([])
            bp.append(["Bidding options", "When to use", "Needs conversion tracking?", "Note"])
            for o in bid["options"]:
                bp.append([o.get("name"), o.get("when"),
                           "Yes" if o.get("needs_tracking") else "No", o.get("note")])
        if bid.get("guardrails"):
            bp.append([])
            bp.append(["Guardrails (avoid overspend)"])
            for g in bid["guardrails"]:
                bp.append(["•", g])
        dev = plan.get("device") or {}
        if dev:
            bp.append(["Device:", dev.get("recommendation")])

        # Seasonality (real Keyword Planner month-on-month) + monthly pacing
        se = wb.create_sheet("Seasonality & Pacing")
        _header(se, ["Month", "Searches (Keyword Planner)", "Index (1.0=avg)", "Demand",
                     "Suggested Budget", "Focus"])
        pacing_by_m = {p["month"]: p for p in plan.get("monthly_pacing", [])}
        for mo in seasonality.get("months", []):
            pm = pacing_by_m.get(mo["month"], {})
            se.append([mo.get("name"), mo.get("searches"), mo.get("index"), mo.get("level"),
                       pm.get("budget"), mo.get("emphasis")])
        if not seasonality.get("available"):
            se.append(["(Keyword Planner seasonality unavailable — budget paced evenly.)"])

    # ---- CPL Optimizer sheet (target CPL → required conversion rate + playbook) ----
    cplp = (plan or {}).get("cpl_plan") or {}
    if cplp:
        cs = wb.create_sheet("CPL Optimizer")
        cs.append([f"Target CPL: ₹{cplp.get('target_cpl_low')}–{cplp.get('target_cpl_high')}"])
        cs.append([f"Required conversion rate: {cplp.get('required_cvr_pct')}% "
                   f"(at optimized ₹{cplp.get('optimized_cpc')} CPC)"])
        cs.append([f"Your click→lead rates — avg {cplp.get('current_cvr_avg_pct')}%, "
                   f"best {cplp.get('current_cvr_best_pct')}%"])
        cs.append(["Verdict:", cplp.get("verdict")])
        cs.append([])
        _header(cs, ["Scenario", "CPC", "Click→lead %", "CPL", "Leads (budget)", "Note"])
        for s in cplp.get("scenarios", []):
            cs.append([s.get("name"), s.get("cpc"), s.get("cvr_pct"), s.get("cpl"),
                       s.get("leads"), s.get("note")])
        cs.append([])
        _header(cs, ["Dial", "Lever", "How"])
        for lv in cplp.get("levers", []):
            cs.append([lv.get("dial"), lv.get("lever"), lv.get("detail")])

    # ---- Keyword History sheet ("keep or drop last time's keywords?") ----
    kh = (gen.scores or {}).get("keyword_history") or {}
    if kh.get("available"):
        rows = kh.get("keywords", [])
        # Union of months across all keywords → one clicks-column per month (MoM pivot).
        months: list[str] = sorted({m["month"] for r in rows for m in r.get("months", [])})
        khs = wb.create_sheet("Keyword History")
        base_cols = ["Keyword", "In Plan?", "Verdict", "Why", "Trend", "Total Clicks",
                     "Total Cost", "Total Conv", "Avg CTR", "Avg CPC", "Avg QS"]
        _header(khs, base_cols + [f"Clicks {m}" for m in months])
        for r in rows:
            by_m = {m["month"]: m for m in r.get("months", [])}
            khs.append(
                [
                    r.get("keyword"),
                    "Yes" if r.get("in_plan") else "No",
                    (r.get("verdict") or "").upper(),
                    r.get("verdict_reason"),
                    r.get("trend"),
                    r.get("total_clicks"),
                    r.get("total_cost"),
                    r.get("total_conversions"),
                    round((r.get("avg_ctr") or 0) * 100, 1) if r.get("avg_ctr") else None,
                    r.get("avg_cpc"),
                    r.get("avg_quality_score"),
                ]
                + [(by_m.get(m, {}).get("clicks") if m in by_m else "") for m in months]
            )
        khs.append([])
        s = kh.get("summary", {})
        khs.append([f"Summary: {s.get('keep', 0)} keep · {s.get('review', 0)} review · "
                    f"{s.get('drop', 0)} drop · {s.get('new', 0)} new keyword(s) in plan."])
        khs.append([f"History window: {kh.get('month_range')} "
                    f"({kh.get('months_covered')} months, campus-scoped real data)."])
        if not kh.get("has_conversions"):
            khs.append(["Note: 0 conversions tracked for this campus — verdicts use "
                        "clicks, CTR, cost and Quality Score, not conversions."])
        new_kw = kh.get("new_in_plan", [])
        if new_kw:
            khs.append([])
            khs.append(["New keywords in this plan (no prior history — no apples-to-apples):"])
            for k in new_kw:
                khs.append([k])

    # ---- Negative Keywords sheet (campus-specific, data-backed) ----
    neg = (gen.scores or {}).get("negative_keywords_detail") or {}
    if neg:
        ns = wb.create_sheet("Negative Keywords")
        ns.append([neg.get("note", "")])
        ns.append([])
        dd = neg.get("from_search_terms", [])
        if dd:
            _header(ns, ["Wasteful search term (from YOUR data)", "Clicks", "Impressions",
                         "Wasted ₹", "Why block it"])
            for d in dd:
                ns.append([d.get("term"), d.get("clicks"), d.get("impressions"),
                           d.get("cost"), d.get("reason")])
            ns.append([f"Total wasted: ₹{neg.get('wasted_spend', 0):,.0f}"])
            ns.append([])
        ns.append(["Preventive negatives (add as broad negatives — block these classes):"])
        for w in neg.get("preventive", []):
            ns.append([w])

    # ---- Landing Page Quality sheet (score + specific fixes) ----
    lq = (gen.scores or {}).get("landing_quality") or {}
    if lq.get("available"):
        ls = wb.create_sheet("Landing Quality")
        ls.append([f"Landing page score: {lq.get('score')}/100 (Grade {lq.get('grade')})"])
        ls.append([])
        _header(ls, ["Check", "On the page?"])
        for c in lq.get("checks", []):
            ls.append([c.get("item"), "Yes" if c.get("ok") else "NO — fix"])
        ls.append([])
        ls.append(["Specific fixes to raise conversion (ranked by impact):"])
        for s in lq.get("suggestions", []):
            ls.append([s])

        au = (gen.scores or {}).get("landing_audit") or {}
        if au.get("available"):
            ls.append([])
            ls.append([f"AUDIT — {au.get('lp_type_label')}"])
            vr = au.get("verdict") or {}
            ls.append([f"Verdict: {vr.get('label')}", vr.get("reason")])
            ls.append([])
            _header(ls, ["Tracking / measurement", "Status", "Where to place it"])
            for c in au.get("tracking_checks", []):
                ls.append([c.get("item"),
                           "On page" if c.get("status") == "present" else "ADD IT",
                           c.get("guidance")])
            ls.append([])
            ls.append(["Retargeting:", au.get("retargeting")])
            ls.append(["Audience segmentation:"])
            for seg in au.get("segmentation", []):
                ls.append(["", seg])

    # ---- Campaign Setup Guide sheet (build-from-scratch checklist) ----
    sg = (gen.scores or {}).get("setup_guide") or {}
    if sg.get("steps"):
        gs = wb.create_sheet("Setup Guide")
        gs.append([f"How to build: {sg.get('campaign_name', '')}"])
        gs.append([f"{sg.get('ready_count', 0)} steps ready · "
                   f"{sg.get('action_count', 0)} need your action"])
        gs.append([])
        _header(gs, ["#", "Step", "Status", "What to do"])
        for i, s in enumerate(sg["steps"], 1):
            gs.append([i, s.get("step"), (s.get("status") or "").upper(), s.get("detail")])

    # widen text columns a little
    for sheet in wb.worksheets:
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 40
        if sheet.max_column >= 5:
            sheet.column_dimensions["E"].width = 50

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
